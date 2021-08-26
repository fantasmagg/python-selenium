# -*- codigo: uft-8 -*-
import datetime
import os
import re
import time
import unittest

import allure
import psycopg2
import openpyxl
import pytest
from behave.model import Scenario
from selenium.common.exceptions import *
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.wait import WebDriverWait
import json
from functions.inicializar import Inicializar
from selenium import webdriver
from selenium.webdriver.ie.options import DesiredCapabilities
from selenium.webdriver.chrome.options import Options as OpcionesChrome
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
Scenarios={}
class Functions(Inicializar):
    #variables para la database
    conexion = psycopg2.connect(user='postgres',
                                password='ADMIN',
                                host='127.0.0.1',
                                port='5432',
                                database='automatizasion')


    #aqui se le vanta el navegador

    def abrir_navegador(self,URL=Inicializar.URL,navegador=Inicializar.NAVEGADOR):
        print("ruta base"+Inicializar.basedir)
        self.ventanas={}
        print("-------------------- ---")
        print(navegador)
        print("-----------------------")

        if navegador ==("IEplorer"):
            caps= DesiredCapabilities.INTERNETEXPLORER.copy()
            #estas son capabilitis ,es como para pasarle una configurasion revia al driver
            caps["platform"]="WINDOWS"
            caps["browserName"]="internet explorer"
            caps["ignoreZoomSetting"]=True
            caps["requireWindowFocus"]=True
            caps["nativeEvents"]=True
            self.driver = webdriver.Ie(Inicializar.basedir+"\\drivers\\IEDriverServer.exe",caps)
            self.driver.implicitly_wait(10)
            self.driver.maximize_window()
            self.driver.get(URL)
            #manejos de ventana
            self.principal = self.driver.window_handles[0]
            self.ventanas = {'Principal':self.driver.window_handles[0]}
            #-----
            self.nWindows = 0
            print(self.ventanas)
            return self.driver

        if navegador == ("CHROME"):
            # estas son capabilitis ,es como para pasarle una configurasion revia al driver
            options = OpcionesChrome()
            options.add_argument('start-maximized')

            self.driver = webdriver.Chrome(options=options,executable_path=Inicializar.basedir+"\\drivers\\chromedriver.exe")
            self.driver.implicitly_wait(10)
            self.driver.get(URL)
            self.principal = self.driver.window_handles[0]
            self.ventanas= {'Principal':self.driver.window_handles[0]}
            print(self.ventanas)
            return self.driver
        if navegador ==("FIREFOX"):
            self.driver = webdriver.Firefox()
            self.driver.implicitly_wait(10)
            self.driver.maximize_window()

            self.driver.get(URL)
            self.principal = self.driver.window_handles[0]
            self.ventanas= {'Principal':self.driver.window_handles[0]}
            return  self.driver

    def tearDown(self):
        print("se cerro el driver")
        self.driver.quit()

 # ----------------------------------------------------------------------------------------------
    #este es una forma para no usar el json , aqui lo ase mos mas directo por eso es hay mas lineas


    #aqui es para localizar el elemento y esas cosas
    def xpath_element(self,XPATH):
        elements = self.driver.find_element_by_xpath(XPATH)
        print("Xpath_Element: se interactua con el elemento "+XPATH)
        return elements
    # aqui esta lo que hace es comprobar si el elemento esta visible y si se le puede hacer click
    def _xpath_element (self,XPATH):
        try:
            wait= WebDriverWait(self.driver,20)
            wait.until(EC.visibility_of_element_located((By.XPATH,XPATH)))
            #wait.until(EC.element_to_be_clickable((By.XPATH)))
            elements = self.driver.find_element_by_xpath(XPATH)
            print("Esperar_Element: Se visualizo el elemeto "+XPATH)
            return elements

        except TimeoutException:
            print(u"Esperar_Element: No presente "+XPATH)
            Functions.tearDown(self)
        except NoSuchElementException:
            print(u"Esperar_Element: No presente " + XPATH)
            Functions.tearDown(self)

    # aqui es para localizar el elemento y esas cosas
    def id_element(self,ID):
        elements = self.driver.find_element_by_id(ID)
        print("id_Element: se interactua con el elemento "+ID)
        return elements

    # aqui esta lo que hace es comprobar si el elemento esta visible y si se le puede hacer click
    def _id_element(self,ID):
        try:
            wait= WebDriverWait(self.driver,20)
            wait.until(EC.visibility_of_element_located((By.ID,ID)))
            #wait.until(EC.element_to_be_clickable((By.ID)))
            elements = self.driver.find_element_by_id(ID)
            print("Esperar_Element: Se visualizo el elemeto "+ID)
            return elements

        except TimeoutException:
            print(u"Esperar_Element: No presente "+ID)
            pytest.skip("valor no encontrado por eso se cancelo la prueva")
            Functions.tearDown(self)
        except NoSuchElementException:
            print(u"Esperar_Element: No presente " + ID)
            pytest.skip("valor no encontrado por eso se cancelo la prueva")
            Functions.tearDown(self)

    # aqui es para localizar el elemento y esas cosas
    def name_element(self,NAME):
        elements = self.driver.find_element_by_name(NAME)
        print("name_Element: se interactua con el elemento " + NAME)
        return elements

    # aqui esta lo que hace es comprobar si el elemento esta visible y si se le puede hacer click
    def _name_element(self,NAME):
        try:
            wait= WebDriverWait(self.driver,10)
            wait.until(EC.visibility_of_element_located((By.NAME,NAME)))
            elements = self.driver.find_element_by_name(NAME)
            print("Esperar_Element: Se visualizo el elemeto " + NAME)
            return elements
        except TimeoutException:
            print(u"Esperar_Element: No presente " + NAME)
            pytest.skip("valor no encontrado por eso se cancelo la prueva")
            Functions.tearDown(self)
        except NoSuchElementException:
            print(u"Esperar_Element: No presente " + NAME)
            Functions.tearDown(self)

#se termina aqui ya hice un ejemeplo usando estas funciones de lecturas es el (hhhh.py)
#----------------------------------------------------------------------------------------------



    #lectura apertura del json

    def get_json_file(self,file):
        json_path= Inicializar.Json+"/"+file +'.json'
        try:
            #esa ("r") ase referencia que lo vamos a solo leer el json
            with open(json_path,"r")as  read_file:
                self.json_strings = json.loads(read_file.read())
                print("get_json_file "+json_path)

                return self.json_strings

        except FileNotFoundError:
            self.json_strings=False
            pytest.skip(u"get_json_file: No se encontro el archivo "+file)
            Functions.tearDown(self)

    def get_entity(self,entity):
        if self.json_strings is False:
            print("Defineel el json para esta prueba")
        else:
            try:
                self.json_ValueToFind = self.json_strings[entity]["ValueToFind"]
                self.json_GetFieldBy = self.json_strings[entity]["GetFieldBy"]
                return True
            except KeyError:
                pytest.skip(u"get_entity no se encontro la entity , esta es a la que nos hizo referensia: " + entity)
                Functions.tearDown(self)
                return None



    def get_elements(self,entity,MytestElement = None):

        Get_Entity = Functions.get_entity(self,entity)
        if Get_Entity is None:
            print("No se encontro el valor en el entity en el json")
        else:
            try:
                if self.json_GetFieldBy.lower()=="id":
                    elements = self.driver.find_element_by_id(self.json_ValueToFind)

                if self.json_GetFieldBy.lower() == "name":
                    elements = self.driver.find_element_by_name(self.json_ValueToFind)

                if self.json_GetFieldBy.lower() == "xpath":
                    if MytestElement is not None:
                        self.json_ValueToFind= self.json_ValueToFind.format(MytestElement)
                        print(self.json_ValueToFind)
                    elements = self.driver.find_element_by_xpath(self.json_ValueToFind)

                if self.json_GetFieldBy.lower() == "link":
                    elements = self.driver.find_element_by_link_text(self.json_ValueToFind)

                if self.json_GetFieldBy.lower() == "css":
                  elements = self.driver.find_element_by_css_selector(self.json_ValueToFind)

                print("get_elements: "+self.json_ValueToFind)
                return elements

            except NoSuchElementException:
                print("get_text No se encuentra el elemento "+ self.json_ValueToFind)
                Functions.tearDown(self)

            except TimeoutException:
                print("get_text el tiempo de espera por elemeneto ya se paso " +self.json_ValueToFind)
                Functions.tearDown(self)



#scrol con js---------------------------------------

    def scrol_js(self, localisador):

        Get_Entity = Functions.get_entity(self, localisador)
        if Get_Entity is None:
            print("No se encontro el valor en el entity en el json")
        else:
            try:
                if self.json_GetFieldBy.lower() == "id":
                    localisador = self.driver.find_element(By.ID,self.json_ValueToFind)
                    self.driver.execute_script("arguments[0].scrollIntoView();", localisador)
                    print(u"scrol_js " + str(localisador))
                    return True
                if self.json_GetFieldBy.lower() == "name":
                    localisador = self.driver.find_element_by_name(self.json_ValueToFind)
                    self.driver.execute_script("arguments[0].scrollIntoView();", localisador)
                    print(u"scrol_js " + localisador)
                    return True

                if self.json_GetFieldBy.lower() == "xpath":
                    localisador = self.driver.find_element(By.XPATH,self.json_ValueToFind)
                    self.driver.execute_script("arguments[0].scrollIntoView();",localisador)
                    print(u"scrol_js " + str(localisador))
                    return True
                if self.json_GetFieldBy.lower() == "link":
                    localisador = self.driver.find_element_by_link_text(self.json_ValueToFind)
                    self.driver.execute_script("arguments[0].scrollIntoView();", localisador)
                    print(u"scrol_js " + localisador)
                    return True

                if self.json_GetFieldBy.lower() == "css":
                    localisador = self.driver.find_element_by_css_selector(self.json_ValueToFind)
                    self.driver.execute_script("arguments[0].scrollIntoView();", localisador)
                    print(u"scrol_js " + localisador)
                    return True

            except NoSuchElementException:
                print("get_text No se encuentra el elemento " + self.json_ValueToFind)
                Functions.tearDown(self)

            except TimeoutException:
                print("get_text el tiempo de espera por elemeneto ya se paso " + self.json_ValueToFind)
                Functions.tearDown(self)








    #esta es para optener el texto de un elemento
    def get_text(self, entity, MyTextElement=None):
        Get_Entity = Functions.get_entity(self, entity)

        if Get_Entity is None:
            print("No se encontro el valor en el Json definido")
        else:
            try:
                if self.json_GetFieldBy.lower() == "id":
                    elements = self.driver.find_element_by_id(self.json_ValueToFind)

                if self.json_GetFieldBy.lower() == "name":
                    elements = self.driver.find_element_by_name(self.json_ValueToFind)

                if self.json_GetFieldBy.lower() == "xpath":
                    if MyTextElement is not None:
                        self.json_ValueToFind = self.json_ValueToFind.format(MyTextElement)
                        print(self.json_ValueToFind)
                    elements = self.driver.find_element_by_xpath(self.json_ValueToFind)

                if self.json_GetFieldBy.lower() == "link":
                    elements = self.driver.find_element_by_partial_link_text(self.json_ValueToFind)

                if self.json_GetFieldBy.lower() == "css":
                    elements = self.driver.find_element_by_css_selector(self.json_ValueToFind)

                print("get_text: " + self.json_ValueToFind)
                print("Text Value : " + elements.text)
                return elements.text

            except NoSuchElementException:
                print("get_text: No se encontró el elemento: " + self.json_ValueToFind)
                Functions.tearDown(self)
            except TimeoutException:
                print("get_text: No se encontró el elemento: " + self.json_ValueToFind)
                Functions.tearDown(self)

    #esta nos ayuda para preguntar si elemento esta visible y si es clicleable
    def espera_el_elemento(self,locator,MytestElement = None):

        Get_Entity = Functions.get_entity(self,locator)
        if Get_Entity is None:
            print("No se encontro el valor en el entity en el json")
        else:
            try:
                if self.json_GetFieldBy.lower()=="id":
                   Wait= WebDriverWait(self.driver,5)
                   Wait.until(EC.visibility_of_element_located((By.ID,self.json_ValueToFind)))
                  # Wait.until(EC.element_to_be_clickable((By.ID,self.json_ValueToFind)))
                   print("Esperar_elemento: se visualize el elemento "+locator)


                if self.json_GetFieldBy.lower() == "name":
                    Wait = WebDriverWait(self.driver, 5)
                    Wait.until(EC.visibility_of_element_located((By.NAME, self.json_ValueToFind)))
                    # Wait.until(EC.element_to_be_clickable((By.NAME,self.json_ValueToFind)))
                    print("Esperar_elemento: se visualize el elemento " + locator)

                if self.json_GetFieldBy.lower() == "xpath":
                    Wait = WebDriverWait(self.driver, 5)
                    if MytestElement is not None:
                        self.json_ValueToFind= self.json_ValueToFind.format(MytestElement)
                        print(self.json_ValueToFind)
                        elements = self.driver.find_element_by_xpath(self.json_ValueToFind)

                    Wait.until(EC.visibility_of_element_located((By.XPATH, self.json_ValueToFind)))
                    # Wait.until(EC.element_to_be_clickable((By.XPATH,self.json_ValueToFind)))
                    print("Esperar_elemento: se visualize el elemento " + locator)

                if self.json_GetFieldBy.lower() == "link":
                    Wait = WebDriverWait(self.driver, 5)
                    Wait.until(EC.visibility_of_element_located((By.LINK_TEXT, self.json_ValueToFind)))
                    # Wait.until(EC.element_to_be_clickable((By.NAME,self.json_ValueToFind)))
                    print("Esperar_elemento: se visualize el elemento " + locator)

                if self.json_GetFieldBy.lower() == "css":
                    Wait = WebDriverWait(self.driver, 5)
                    Wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, self.json_ValueToFind)))
                    # Wait.until(EC.element_to_be_clickable((By.NAME,self.json_ValueToFind)))
                    print("Esperar_elemento: se visualize el elemento " + locator)
                    return True

            except NoSuchElementException:
                 print("get_text No se encuentra el elemento "+ self.json_ValueToFind)
                 Functions.tearDown(self)

            except TimeoutException:
                 print("get_text el tiempo de espera por elemeneto ya se paso " +self.json_ValueToFind)
                 Functions.tearDown(self)



    def get_select_element(self, entity):

        Get_Entity = Functions.get_entity(self, entity)
        if Get_Entity is None:
            print("No se encontro el valor en el entity en el json")
        else:
            try:
                if self.json_GetFieldBy.lower() == "id":
                   select = Select(self.driver.find_element_by_id(self.json_ValueToFind))

                if self.json_GetFieldBy.lower() == "name":
                    select = Select(self.driver.find_element_by_name(self.json_ValueToFind))

                if self.json_GetFieldBy.lower() == "xpath":
                    select = Select(self.driver.find_element_by_xpath(self.json_ValueToFind))

                if self.json_GetFieldBy.lower() == "link":
                    select = Select(self.driver.find_element_by_link_text(self.json_ValueToFind))

                if self.json_GetFieldBy.lower() == "css":
                    select = Select(self.driver.find_element_by_css_selector(self.json_ValueToFind))

                print("get_select: " + self.json_ValueToFind)

                return select
                # se puede selesionar por texto visible o por value
            except NoSuchElementException:
                print("get_text No se encuentra el elemento " + self.json_ValueToFind)
                Functions.tearDown(self)

            except TimeoutException:
                print("get_text el tiempo de espera por elemeneto ya se paso " + self.json_ValueToFind)
                Functions.tearDown(self)

    def selector_text(self, entity,text):
        Functions.get_select_element(self,entity).select_by_visible_text(text)

#para salir
    def switch_to_parent_frame(self):
        self.driver.switch_to.parent_frame()
#para entrar
    def switch_to_iframe(self,localtor):
        iframe = Functions.get_elements(self,localtor)
        self.driver.switch_to.frame(iframe)


    def send_key_text(self,entity,texto):
        Functions.get_elements(self,entity).clear()
        Functions.get_elements(self,entity).send_keys(texto)

    #estas funciones son para poder pasar entre ventanas

    def switch_to_windows_name(self,ventana):
        if ventana in self.ventanas:
            self.driver.switch_to.window(self.ventanas[ventana])
            Functions.page_has_loaded(self)
            print("volviendo a "+ventana+" : "+self.ventanas[ventana])
        else:
            self.nWindows=len(self.driver.window_handles)-1
            self.ventanas[ventana]=self.driver.window_handles[int(self.nWindows)]
            self.driver.switch_to.window(self.ventanas[ventana])
            self.driver.maximize_window()
            print(self.ventanas)
            print("Estas en "+ventana +" : "+self.ventanas[ventana])
            Functions.page_has_loaded(self)


    def new_window(self,URL):
        #esto dise que si le pasamos una URL la abra en otra ventana
        self.driver.execute_script(f'''window.open("{URL}","_blank");''')
        Functions.page_has_loaded(self)


    def page_has_loaded(self):
        driver= self.driver
        print("Checking if {} page is loaded , ".format(self.driver.current_url))
        page_state= driver.execute_script('return document.readyState;')
        yield
        WebDriverWait(driver,10).until(lambda driver: page_state=="complete")
        assert page_state == 'complete',"no se a cargado completamente la pagina "





    def click_js(self, localisador):

        Get_Entity = Functions.get_entity(self, localisador)
        if Get_Entity is None:
            print("No se encontro el valor en el entity en el json")
        else:
            try:
                if self.json_GetFieldBy.lower() == "id":
                    localisador = self.driver.find_element_by_id(self.json_ValueToFind)
                    self.driver.execute_script("arguments[0].click();", localisador)
                    print(u"scrol_js " + str(localisador))
                    return True
                if self.json_GetFieldBy.lower() == "name":
                    localisador = self.driver.find_element_by_name(self.json_ValueToFind)
                    self.driver.execute_script("arguments[0].click();", localisador)
                    print(u"scrol_js " + localisador)
                    return True

                if self.json_GetFieldBy.lower() == "xpath":
                    localisador = self.driver.find_element_by_xpath(self.json_ValueToFind)
                    self.driver.execute_script("arguments[0].click();",localisador)
                    print(u"scrol_js " + str(localisador))
                    return True
                if self.json_GetFieldBy.lower() == "link":
                    localisador = self.driver.find_element_by_link_text(self.json_ValueToFind)
                    self.driver.execute_script("arguments[0].click();", localisador)
                    print(u"scrol_js " + localisador)
                    return True

                if self.json_GetFieldBy.lower() == "css":
                    localisador = self.driver.find_element_by_css_selector(self.json_ValueToFind)
                    self.driver.execute_script("arguments[0].click();", localisador)
                    print(u"click_js " + localisador)
                    return True

            except NoSuchElementException:
                print("get_text No se encuentra el elemento " + self.json_ValueToFind)
                Functions.tearDown(self)

            except TimeoutException:
                print("get_text el tiempo de espera por elemeneto ya se paso " + self.json_ValueToFind)
                Functions.tearDown(self)






    def esperar(self , timeload):
        print("esperar: Inicia ("+(str(timeload)+")"))
        try:
                totalWait=0
                while(totalWait < timeload):
                    time.sleep(1)
                    totalWait=totalWait +1;
                    print(totalWait)
        finally:
            print("Eperar carga finalisada...")
#-------------------------------------------------------
    def alert_windows(self,accept="accept"):
        try:
            wait = WebDriverWait(self.driver,30)
            wait.until(EC.alert_is_present(),print("Esperado alerta..."))

            alert = self.driver.switch_to.alert

            print(alert.text)
            if accept.lower()=="accept":
                alert.accept()
                print("Click in accept")
            else:
                alert.dismiss()
                print("click in dismiss")

        except NoAlertPresentException:
            print('alerta no presente')
        except NoSuchWindowException:
            print('alerta no presente')
        except TimeoutException:
            print('alerta no presente')


    def send_especific_keys(self, elemet,key):
        if key == 'Enter':
            Functions.get_elements(self,elemet).send_keys(Keys.ENTER)
        if key =='Tab':
            Functions.get_elements(self,elemet).send_keys(Keys.TAB)
        if key == 'Space':
            Functions.get_elements(self,elemet).send_keys(Keys.SPACE)
        time.sleep(3)

    def send_key_text (self , entity,text):
        Functions.get_elements(self,entity).clear()
        Functions.get_elements(self,entity).send_keys(text)


    def assert_text(self,elemento,texto):
        Get_Entity= Functions.get_entity(self,elemento)
        if Get_Entity is None:
            print("No se encontro el valor en el entity en el json")
        else:

                if self.json_GetFieldBy.lower() == "id":

                    Wait = WebDriverWait(self.driver, 5)
                    Wait.until(EC.visibility_of_element_located((By.ID, self.json_ValueToFind)))
                    ObbteneTexto = self.driver.find_element_by_id(self.json_ValueToFind).text

                if self.json_GetFieldBy.lower() == "name":

                    Wait = WebDriverWait(self.driver, 5)
                    Wait.until(EC.visibility_of_element_located((By.NAME, self.json_ValueToFind)))
                    ObbteneTexto = self.driver.find_element_by_name(self.json_ValueToFind).text

                if self.json_GetFieldBy.lower() == "xpath":

                   Wait = WebDriverWait(self.driver, 5)
                   Wait.until(EC.visibility_of_element_located((By.XPATH, self.json_ValueToFind)))
                   ObbteneTexto = self.driver.find_element_by_xpath(self.json_ValueToFind).text


                if self.json_GetFieldBy.lower() == "link":

                    Wait = WebDriverWait(self.driver, 5)
                    Wait.until(EC.visibility_of_element_located((By.LINK_TEXT, self.json_ValueToFind)))
                    ObbteneTexto = self.driver.find_element_by_link_text(self.json_ValueToFind).text



                if self.json_GetFieldBy.lower() == "css":

                    Wait = WebDriverWait(self.driver, 5)
                    Wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, self.json_ValueToFind)))
                    ObbteneTexto = self.driver.find_element_by_css_selector(self.json_ValueToFind).text


        print("Verificar texto el valor mostrado en: "+elemento)
        print(" es "+ObbteneTexto+" el esperado es : "+texto)
        assert texto == texto,"no son iguales"

    def check_text(self, elemento):
        Get_Entity = Functions.get_entity(self, elemento)
        if Get_Entity is None:
            print("No se encontro el valor en el entity en el json")
        else:
            try:
                if self.json_GetFieldBy.lower() == "id":
                    Wait = WebDriverWait(self.driver, 5)
                    Wait.until(EC.visibility_of_element_located((By.ID, self.json_ValueToFind)))
                    print("se encontro el elemento : " + elemento)
                    return True

                if self.json_GetFieldBy.lower() == "name":
                    Wait = WebDriverWait(self.driver, 5)
                    Wait.until(EC.visibility_of_element_located((By.NAME, self.json_ValueToFind)))
                    print("se encontro el elemento : " + elemento)
                    return True

                if self.json_GetFieldBy.lower() == "xpath":
                    Wait = WebDriverWait(self.driver, 5)
                    Wait.until(EC.visibility_of_element_located((By.XPATH, self.json_ValueToFind)))
                    print("se encontro el elemento : " + elemento)
                    return True

                if self.json_GetFieldBy.lower() == "link":
                    Wait = WebDriverWait(self.driver, 5)
                    Wait.until(EC.visibility_of_element_located((By.LINK_TEXT, self.json_ValueToFind)))
                    print("se encontro el elemento : " + elemento)
                    return True

                if self.json_GetFieldBy.lower() == "css":
                    Wait = WebDriverWait(self.driver, 5)
                    Wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, self.json_ValueToFind)))
                    print("se encontro el elemento : " + elemento)
                    return True
            except NoAlertPresentException:
                print('alerta no presente')
            except NoSuchWindowException:
                print('alerta no presente')
            except TimeoutException:
                print('se agoto el tiempo de espera')


#variable esenarios

    def create_variable_scenary (self,key,value):
        Scenario[key]=value
        print(Scenario)
        print("Se almaceno la key"+key+":"+value)


    def save_variable_ecenario(self,elemet, variable):
          Scenarios[variable]=Functions.get_text(self,elemet)
          print(Scenarios)
          print("se alamaceno el valor "+variable+" :"+ Scenarios[variable])

    def get_varible_scenario(self,variable):
        self.variable = Scenarios[variable]
        print(f"get_varible_scenario :{self.variable}")
        return self.variable


    def compare_with_variable(self,elemet,variable):
        variable_scenario = str(Scenarios[variable])
        elemet_text = str(Functions.get_text(self,elemet))
        _exist =(variable_scenario in elemet_text)
        print(f'comprobando las variable .... verificando que si {variable_scenario} esta presente en {elemet_text} : {_exist}')
        assert _exist==True, f'{variable_scenario} != {elemet_text}'



    def textDateEnviaromenReploce(self,text):
        if text == 'today':
            self.today= datetime.date.today()
            text= self.today.strftime(Inicializar.DateFormat)
        if text == 'yesterday':
            self.today = datetime.date.today() - datetime.timedelta(days=1)
            text = self.today.strftime(Inicializar.DateFormat)

        if text == 'last Month':
            self.today= datetime.date.today() - datetime.timedelta(days=30)
            text = self.today.strftime(Inicializar.DateFormat)

        return text

#------------------------------------------------Excel
    def leer_celdaEXl (self ,celda):
        wb = openpyxl.load_workbook(Inicializar.Excel)
        sheet = wb ["DataTest"]
        valor = str(sheet[celda].value)
        print("--------------------------------")
        print("el libro de excel utilisado es de :"+Inicializar.Excel)
        print("se escribio en la celda "+str(celda)+" el valor: "+str(valor))
        print("--------------------------------")

    def leer_celdaEXl(self, celda,valor):
        wb = openpyxl.load_workbook(Inicializar.Excel)
        hoja = wb["DataTest"]
        hoja[celda]= valor
        wb.save(Inicializar.Excel)
        print("--------------------------------")
        print("el libro de excel utilisado es de :" + Inicializar.Excel)
        print("se escribio en la celda " + str(celda) + " el valor: " + str(valor))
        print("--------------------------------")


    #-----------------------------------------------conect bbdd
    # def pyodbc_conn(self,_user=Inicializar.user,_password=Inicializar.password,_host=Inicializar.host,_port=Inicializar.port,_database=Inicializar.database):
    #     try:
    #         config=dict(
    #              usuario = _user,
    #              password = 'ADMIN',
    #              host = '127.0.0.1',
    #              port = '5432',
    #              database = 'test_db'  )
    #         conn_str=(
    #             'SERVER=(usuaro)'+
    #
    #         
    #         )
    #---------------------------------------------------------------------base de datos
    @classmethod
    def databaseSelect (cls):
        try:
            with cls.conexion:
                with cls.conexion.cursor() as cursor:
                    # %s eso es para dicerle que el valor sera pasado por un avariable
                    id_persona = 3
                    # setencia = 'SELECT * FROM personas WHERE id_persona =%s'
                    setencia = 'SELECT * FROM pythondata'
                    # id_persona lo tuvimos que convertir en una tupla (id_persona,)
                    # cursor.execute(setencia,(id_persona,))
                    cursor.execute(setencia)
                    # cursor.fetchall() con esa instrusion es que podemos sacar los valores de execute (setencia)
                    resgitro = cursor.fetchall()
                    # o
                    # .fetchone() es para que solo nostraiga un registro o usuario
                    # resgitro = cursor.fetchone()
                    listaDatos=[]
                    for resgitros in resgitro:
                        dato1 = (resgitros[0],resgitros[1])
                        listaDatos.append(dato1)
                    print(listaDatos)
                    return listaDatos

        except Exception as e:
            print(f" ocurrio un error {e}")



    @classmethod
    def insertarValores(cls,datos):
        try:
            with cls.conexion:
                with cls.conexion.cursor() as cursor:
                    sentencia = 'INSERT INTO pythondata(datos)values(%s)'

                    valores = (datos,)
                    cursor.execute(sentencia, valores)
                    registrosINSERT = cursor.rowcount
                    print(f'Registros insertados {registrosINSERT}')
        except Exception as e:
            print(f" ocurrio un error {e}")

    @classmethod
    def cerrarConesion(cls):
        cls.conexion.close()
        print('se cerro la conexion')

#-----------------------------------------------------------------------------Excel
    def leer_celda(self,celda):
        wb = openpyxl.load_workbook(Inicializar.Excel)

        #esto es la hoja
        sheet = wb['DataTest']
        valor = str(sheet[celda].value)
        print(u'------------------------------------------------------------')
        print(u'El libro de excel utilizado es :'+Inicializar.Excel)
        print(u"el valor de la celda es:"+valor)
        print(u'------------------------------------------------------------')
        return valor

    def escribir_celda(self,celda,valor):
        wb = openpyxl.load_workbook(Inicializar.Excel)
        hoja = wb ["DataTest"]
        hoja[celda]=valor
        wb.save(Inicializar.Excel)
        print(u'------------------------------------')
        print(u'El libro de excel utilizado es :' + Inicializar.Excel)
        print(f'se Escribio en la celda :{str(celda)}  el valor {str(valor)}')
        print(u'------------------------------------')

    #--------------------------------------------captura de pantalla-----------------------------
    def hora_Actual(self):
        self.hora = time.strftime(Inicializar.HourFormat)  # FORmato 24 horas
        return self.hora
    def crear_path(self):

        dia = time.strftime("%d-%m-%Y") #formato aaaa/m/dd

        GeneralPath = Inicializar.Path_Evidencias
        DriverTest = Inicializar.NAVEGADOR
        TestCase = self.__class__.__name__
        horaAct = Functions.hora_Actual(self)
        x = re.search("Conetext",TestCase)
        if (x):
            path = f"{GeneralPath}/{dia}/{DriverTest}/{horaAct}/"

        else:
            path = f"{GeneralPath}/{dia}/{TestCase}/{DriverTest}/{horaAct}/"
        if not os.path.exists(path): #si no existe el diretorio lo crea
            os.makedirs(path)

        return path

    def captura_pantalla(self, TestCase=' captura '):
        PATH = Functions.crear_path(self)
        img = f'{PATH}/{TestCase}_({str(Functions.hora_Actual(self))}).png'
        self.driver.get_screenshot_as_file(img)
        print(img)
        return img

    def captura(self,Descripcion):
        allure.attach(self.driver.get_screenshot_as_png(),Descripcion,attachment_type=allure.attachment_type.PNG)




    #---------------------------------------------

    # esta nos ayuda para preguntar si elemento esta visible y si es clicleable
    def validar_elemento(self, locator):
        TIME_OUT =10
        texto="Esperar_elemento: se visualize el elemento "
        Get_Entity = Functions.get_entity(self, locator)
        if Get_Entity is None:
            print("No se encontro el valor en el entity en el json")
        else:
            try:
                if self.json_GetFieldBy.lower() == "id":
                    Wait = WebDriverWait(self.driver, TIME_OUT)
                    Wait.until(EC.visibility_of_element_located((By.ID, self.json_ValueToFind)))
                    Wait.until(EC.element_to_be_clickable((By.ID,self.json_ValueToFind)))
                    print(texto+" "+ locator)
                    return True

                if self.json_GetFieldBy.lower() == "name":
                    Wait = WebDriverWait(self.driver, TIME_OUT)
                    Wait.until(EC.visibility_of_element_located((By.NAME, self.json_ValueToFind)))
                    Wait.until(EC.element_to_be_clickable((By.NAME,self.json_ValueToFind)))
                    print(texto +" " + locator)
                    return True

                if self.json_GetFieldBy.lower() == "xpath":
                    Wait = WebDriverWait(self.driver, TIME_OUT)
                    Wait.until(EC.visibility_of_element_located((By.XPATH, self.json_ValueToFind)))
                    Wait.until(EC.element_to_be_clickable((By.XPATH,self.json_ValueToFind)))
                    print(texto+" " + locator)
                    return True

                if self.json_GetFieldBy.lower() == "link":
                    Wait = WebDriverWait(self.driver, TIME_OUT)
                    Wait.until(EC.visibility_of_element_located((By.LINK_TEXT, self.json_ValueToFind)))
                    Wait.until(EC.element_to_be_clickable((By.NAME,self.json_ValueToFind)))
                    print(texto+" " + locator)
                    return True

                if self.json_GetFieldBy.lower() == "css":
                    Wait = WebDriverWait(self.driver, TIME_OUT)
                    Wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, self.json_ValueToFind)))
                    Wait.until(EC.element_to_be_clickable((By.NAME,self.json_ValueToFind)))
                    print(texto+" " + locator)
                    return True

            except NoSuchElementException:
                print("get_text No se encuentra el elemento " + self.json_ValueToFind)
                return False